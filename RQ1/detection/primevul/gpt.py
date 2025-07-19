



# newmsg
from openai import OpenAI
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import os
# 读取CSV文件
df = pd.read_excel('D)

# 设置OpenAI客户端
client = OpenAI(
    base_url='https:',
    api_key='sk-',
)


# 定义一个函数来分析代码并返回结果
def analyze_code(code, commit_msg):
    code = code[:10000]
    messages = [
        {
            "role": "user",
            "content": "Given a piece of code and commit message, determine whether it is vulnerable or security."
                       "Desired Format:"
                       "is_vul: <1/0>"
                       "code snippet: {}"
                       "commit message: {}".format(code, commit_msg),
        }
    ]

    response = client.chat.completions.create(
        messages=messages,
        model="gpt-3.5-turbo",
        temperature=0.1,
        top_p=1,
    )
    for message in response.choices:
        print(message.message.content)
    result = {}
    for message in response.choices:
        for line in message.message.content.split('\n'):
            if line.startswith('is_vul:'):
                result['is_vul'] = line[len('is_vul: '):].strip()
    return result


def format_row(row):
    return (f""
            f"Summary: {row['Summary']} "
            f"Background: {row['Background']} "
            f"Impact: {row['Impact']} "
            )
msg = [format_row(row) for _, row in df.iterrows()]


# 创建一个新的DataFrame来存储结果
results_df = pd.DataFrame(columns=['cve', 'is_vul'])
count = 0
start_index = 0
csv_filename = 'PrimeVul_gpt_newmsg.xlsx'

# 遍历DataFrame，分析代码并将结果逐条写入Excel文件
for index, row in df.iloc[start_index:].iterrows():
    result = analyze_code(row['func'], msg[index])
    new_row = [
        row['cve'],
        result.get('is_vul', ''),
    ]
    # 如果文件不存在，创建一个新的工作簿并写入表头和数据
    if not os.path.exists(csv_filename):
        wb = Workbook()
        ws = wb.active
        # 写入表头
        ws.append(['cve', 'is_vul'])
        # 写入数据
        ws.append(new_row)
        wb.save(csv_filename)
    else:
        # 如果文件存在，追加数据
        wb = load_workbook(csv_filename)
        ws = wb.active
        ws.append(new_row)
        wb.save(csv_filename)
    count += 1
    print("处理完成{}个啦".format(count))











import pandas as pd

# 读取 Excel 文件
df = pd.read_excel('PrimeVul_gpt_newmsg.xlsx')

# 统计 is_vul 列中值为 1 的数量
count_1 = df['is_vul'].sum()

print(f"is_vul 列中值为 1 的数量为：{count_1}")