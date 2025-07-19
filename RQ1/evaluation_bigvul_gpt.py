from openai import OpenAI
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import os
# 读取CSV文件
df = pd.read_csv('../data/BigVul/BigVul_new_recom.csv')
df_2 = pd.read_excel('../data/BigVul/BigVul_new_com_msg.xlsx')
# 设置OpenAI客户端
client = OpenAI(
    base_url='https://',
    api_key='sk-',
)

# 定义一个函数来分析代码并返回结果
def analyze_code(code, commit_code, commit_msg):
    code = code[:10000]
    commit_code = commit_code[:500]

    messages = [
        {
            "role": "user",
            "content": "Here is a piece of software bug code, the changes made in the corresponding commit and the corresponding commit message."
                       "Please rate each commit message on a scale of 1 to 5, where higher is better."
                       "There are the following indicators: 1. Accurately summarize and submit change information; 2. Natural and concise statements; 3. Help software engineers quickly understand change information."
                       "Be objective, like a software engineering and score a little lower."
                       "Expected output format:"
                       "Commit Message: <your rating>"
                       "Input:"
                       "Code: {}"
                       "Code Change: {}"
                       "Commit message: {}".format( code, commit_code, commit_msg),
        }
    ]

    response = client.chat.completions.create(
        messages=messages,
        model="gpt-3.5-turbo",
        temperature=0.1,
        top_p=1,
    )
    for message in response.choices:
        print(message.message.content)
    result = {}
    for message in response.choices:
        for line in message.message.content.split('\n'):
            if line.startswith('Commit Message:'):
                result['Commit Message'] = line[len('Commit Message: '):].strip()
    return result


# 创建一个新的DataFrame来存储结果
results_df = pd.DataFrame(columns=['CVE ID', 'Commit Message'])
count = 1162
start_index = 1162
csv_filename = 'BigVul_score_gpt.xlsx'

# 遍历DataFrame，分析代码并将结果逐条写入Excel文件
for index, row in df.iloc[start_index:].iterrows():
    # 合并df_2中的四列数据，并添加列名作为前缀
    additional_info = "\n".join(
        [f"{col}: {df_2.loc[index, col]}" for col in ['Summary', 'Background', 'Impact', 'Fix']]
    )
    print(additional_info)  # 打印合并后的信息，用于调试
    result = analyze_code(row['func_before_recom'], row['patch'], additional_info)
    new_row = [
        row['CVE ID'],
        result.get('Commit Message', ''),
    ]
    # 如果文件不存在，创建一个新的工作簿并写入表头和数据
    if not os.path.exists(csv_filename):
        wb = Workbook()
        ws = wb.active
        # 写入表头
        ws.append(['CVE ID', 'Commit Message'])
        # 写入数据
        ws.append(new_row)
        wb.save(csv_filename)
    else:
        # 如果文件存在，追加数据
        wb = load_workbook(csv_filename)
        ws = wb.active
        ws.append(new_row)
        wb.save(csv_filename)
    count += 1
    print("处理完成{}个啦".format(count))



import pandas as pd

# 读取Excel文件
file_path = 'BigVul_score_gpt.xlsx'  # 替换为你的文件路径
df = pd.read_excel(file_path)

# 检查Commit Message列是否存在
if 'Commit Message' not in df.columns:
    print("错误：'Commit Message' 列不存在于Excel文件中！")
else:
    # 初始化异常列表
    anomalies = []

    # 遍历Commit Message列
    for index, value in df['Commit Message'].items():  # 使用items()替代iteritems()
        # 检查是否为空
        if pd.isna(value):
            anomalies.append((index, value, "空值"))
        else:
            # 检查是否为数字
            try:
                num = float(value)
                # 检查是否在1-5之间
                if not 1 <= num <= 5:
                    anomalies.append((index, value, "数字不在1-5之间"))
            except ValueError:
                anomalies.append((index, value, "非数字内容"))

    # 输出异常情况
    if anomalies:
        print("发现异常：")
        for anomaly in anomalies:
            print(f"行号 {anomaly[0]}: 值 {anomaly[1]}，问题：{anomaly[2]}")
            # 将异常值清空
            df.at[anomaly[0], 'Commit Message'] = None
    else:
        print("未发现异常，所有值均符合要求。")

    # 计算剩余正常值的平均数
    normal_values = df['Commit Message']
    if not normal_values.empty:
        average = normal_values.mean()
        print(f"所有值（包括替换后的空值）的平均数为：{average}")
    else:
        print("没有值，无法计算平均数。")