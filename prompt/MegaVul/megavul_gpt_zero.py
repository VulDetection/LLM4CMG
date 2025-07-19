from openai import OpenAI
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import os
# 读取CSV文件
df = pd.read_excel('../../data/MegaVul/MegaVul_test.xlsx')

file = "gpt"
temp = 0.1
model="gpt-3.5-turbo"

csv_filename = 'bigvul_{}_zero.xlsx'.format(file)

# 设置OpenAI客户端
client = OpenAI(
    base_url='https://',
    api_key='sk-',
)

# 定义一个函数来分析代码并返回结果
def analyze_code(code, change):
    code = code[:10000]
    messages = [
        {
            "role": "user",
            "content": "Please generate a short commit message including <Summary>, <Background>, <Impact> and <Fix> for the following function and code change within 80 words."
                       "Desired Format:"
                       "Summary:"
                       "<Module>: <One-line fix summary>"
                       "Background:"
                       "<Briefly explain the root cause>"
                       "Impact:"
                       "<Describe the security impact if exploited>"
                       "Fix:"
                       "<Explain how this patch resolves the problem.>"
                       "Code Snippet: {}"
                       "Code Change: {}".format(code, change),
        }
    ]

    response = client.chat.completions.create(
        messages=messages,
        model = model,
        temperature = temp,
        top_p=1,
    )
    for message in response.choices:
        print(message.message.content)
    result = {}
    for message in response.choices:
        for line in message.message.content.split('\n'):
            line = line.strip().strip('*').strip()
            if line.startswith('Summary:'):
                result['Summary'] = line[len('Summary: '):].strip()
            elif line.startswith('Background:'):
                result['Background'] = line[len('Background: '):].strip()
            elif line.startswith('Impact:'):
                result['Impact'] = line[len('Impact: '):].strip()
            elif line.startswith('Fix:'):
                result['Fix'] = line[len('Fix: '):].strip()
    return result


# 创建一个新的DataFrame来存储结果
results_df = pd.DataFrame(columns=['cve_id', 'Summary', 'Background', 'Impact', 'Fix'])
count = 0
start_index = 0


# 遍历DataFrame，分析代码并将结果逐条写入Excel文件
for index, row in df.iloc[start_index:].iterrows():
    result = analyze_code(row['func_before_recom'], row['diff_func'])
    new_row = [
        row['cve_id'],
        result.get('Summary', ''),
        result.get('Background', ''),
        result.get('Impact', ''),
        result.get('Fix', ''),
    ]
    # 如果文件不存在，创建一个新的工作簿并写入表头和数据
    if not os.path.exists(csv_filename):
        wb = Workbook()
        ws = wb.active
        # 写入表头
        ws.append(['CVE ID', 'Summary', 'Background', 'Impact', 'Fix'])
        # 写入数据
        ws.append(new_row)
        wb.save(csv_filename)
    else:
        # 如果文件存在，追加数据
        wb = load_workbook(csv_filename)
        ws = wb.active
        ws.append(new_row)
        wb.save(csv_filename)
    count += 1
    print("处理完成{}个啦".format(count))




# # cal score
# import pandas as pd
# import torch.nn.functional as F
#
# # 读取预测值（来自 bigvul_gpt.xlsx）
# pred_df = pd.read_excel('bigvul_gpt.xlsx', usecols=['Summary', 'Background', 'Impact', 'Fix'])
#
# # 读取真实值（来自 BigVul_test.csv）
# gt_df = pd.read_excel('../data/BigVul/BigVul_test.xlsx', usecols=['Summary.1', 'Background', 'Impact', 'Fix'])
#
# # 读取真实值（来自 BigVul_test.csv）
# gt_code_df = pd.read_excel('../data/BigVul/BigVul_test.xlsx', usecols=['patch'])
#
# # 辅助函数：将每一行拼接成字符串
# def format_row_pred(row):
#     return f"Summary: {row['Summary']} Background: {row['Background']} Impact: {row['Impact']} Fix: {row['Fix']}"
# # 辅助函数：将每一行拼接成字符串
# def format_row_gt(row):
#     return f"Summary: {row['Summary.1']} Background: {row['Background']} Impact: {row['Impact']} Fix: {row['Fix']}"
# # 构建 pred 和 gt 列表
# pred = [format_row_pred(row) for _, row in pred_df.iterrows()]
# gt = [format_row_gt(row) for _, row in gt_df.iterrows()]
# gt_code = gt_code_df['patch']
#
# from nlgeval import NLGEval
#
# nlgeval = NLGEval(no_skipthoughts=True, no_glove=True)
#
# metric = nlgeval.compute_metrics(ref_list=[gt], hyp_list=pred)
# Bleu_4 = metric['Bleu_4']
# METEOR = metric['METEOR']
# ROUGE_L = metric['ROUGE_L']
# print("Bleu_4: {:.4f}".format(Bleu_4))
# print("METEOR: {:.4f}".format(METEOR))
# print("ROUGE_L: {:.4f}".format(ROUGE_L))
#
#
# from bert_score import score
#
# P, R, F1 = score(pred, gt, lang="en", verbose=True)
# print(f"BERTScore F1: {F1.mean():.4f}")
#
#
# from sentence_transformers import SentenceTransformer, util
#
# model = SentenceTransformer('all-MiniLM-L6-v2')
#
# # 编码句子
# pred_emb = model.encode(pred, convert_to_tensor=True)
# gt_emb = model.encode(gt, convert_to_tensor=True)
#
# # 计算每一对的余弦相似度
# similarities = []
# for i in range(len(pred)):
#     sim = util.cos_sim(pred_emb[i], gt_emb[i]).item()
#     similarities.append(sim)
#
# # 计算平均相似度
# average_similarity = sum(similarities) / len(similarities)
#
# print(f"Sentence-BERT with Cosine Similarity (SBCS): {average_similarity:.4f}")
#
#
#
# # 使用上面编码好的 pred_emb 和 gt_emb
# euc_dist = F.pairwise_distance(pred_emb, gt_emb, p=2)
# # print(f"SBERT Euclidean Distance: {euc_dist}")
# average_euc_dist = euc_dist.mean().item()
# print(f"Sentence-BERT with Euclidean Distance (SBED): {average_euc_dist:.4f}")
#
#
#
# import tensorflow_hub as hub
# import numpy as np
# from sklearn.metrics.pairwise import cosine_similarity
#
# # 加载 USE 模型
# use_model = hub.load("https://tfhub.dev/google/universal-sentence-encoder/4")
#
# # 编码句子
# pred_emb = use_model(pred).numpy()
# gt_emb = use_model(gt).numpy()
#
# # 计算余弦相似度
# cos_sim = cosine_similarity(pred_emb, gt_emb)
# print(f"Universal Sentence Encoder with Cosine Similarity (USECS): {cos_sim[0][0]:.4f}")
#
#
# from transformers import AutoTokenizer, AutoModel
# import torch
# from sklearn.metrics.pairwise import cosine_similarity
#
# # 加载 CodeBERT 模型和 tokenizer
# tokenizer = AutoTokenizer.from_pretrained("D:/sh_2024_2028/研究内容-3/实验/Vul_eval/models/bert")
# model = AutoModel.from_pretrained("D:/sh_2024_2028/研究内容-3/实验/Vul_eval/models/bert")
#
# side_scores = []
#
# for p, g in zip(pred, gt_code):
#     # 编码句子
#     inputs_p = tokenizer(p, return_tensors="pt", truncation=True, padding=True)
#     inputs_g = tokenizer(g, return_tensors="pt", truncation=True, padding=True)
#
#     with torch.no_grad():
#         emb_p = model(**inputs_p).pooler_output.numpy()
#         emb_g = model(**inputs_g).pooler_output.numpy()
#
#     # 计算余弦相似度（即 SIDE 分数）
#     sim = cosine_similarity(emb_p, emb_g)[0][0]
#     side_scores.append(sim)
#
# # 输出平均 SIDE 分数
# average_side = sum(side_scores) / len(side_scores)
# print(f"SIDE: {average_side:.4f}")