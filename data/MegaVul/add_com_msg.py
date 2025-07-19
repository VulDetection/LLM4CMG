from openai import OpenAI
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import os
# 读取CSV文件
df = pd.read_csv('MegaVul_new_recom.csv')

# 设置OpenAI客户端
client = OpenAI(
    base_url='htt###',
    api_key='sk-###',
)

example_code = '''
static int rndis_query_response(USBNetState *s,
                rndis_query_msg_type *buf, unsigned int length)
{
    rndis_query_cmplt_type *resp;
    uint8_t infobuf[sizeof(oid_supported_list)];
    uint32_t bufoffs, buflen;
    int infobuflen;
    unsigned int resplen;
    bufoffs = le32_to_cpu(buf->InformationBufferOffset) + 8;
    buflen = le32_to_cpu(buf->InformationBufferLength);
    if (bufoffs + buflen > length)
        return USB_RET_STALL;
    infobuflen = ndis_query(s, le32_to_cpu(buf->OID),
                            bufoffs + (uint8_t *) buf, buflen, infobuf,
                            sizeof(infobuf));
    resplen = sizeof(rndis_query_cmplt_type) +
            ((infobuflen < 0) ? 0 : infobuflen);
    resp = rndis_queue_response(s, resplen);
    if (!resp)
        return USB_RET_STALL;
    resp->MessageType = cpu_to_le32(RNDIS_QUERY_CMPLT);
    resp->RequestID = buf->RequestID;
    resp->MessageLength = cpu_to_le32(resplen);
    if (infobuflen < 0) {
        resp->Status = cpu_to_le32(RNDIS_STATUS_NOT_SUPPORTED);
        resp->InformationBufferLength = cpu_to_le32(0);
        resp->InformationBufferOffset = cpu_to_le32(0);
        return 0;
    }
    resp->Status = cpu_to_le32(RNDIS_STATUS_SUCCESS);
    resp->InformationBufferOffset =
            cpu_to_le32(infobuflen ? sizeof(rndis_query_cmplt_type) - 8 : 0);
    resp->InformationBufferLength = cpu_to_le32(infobuflen);
    memcpy(resp + 1, infobuf, infobuflen);
    return 0;
}
'''


example_commit = '''
usb: check RNDIS buffer offsets & length

When processing remote NDIS control message packets,
the USB Net device emulator uses a fixed length(4096) data buffer.
The incoming informationBufferOffset & Length combination could
overflow and cross that range. Check control message buffer
offsets and length to avoid it.
'''

example_out = '''
Summary: usb: Prevent out-of-bounds access in RNDIS control message handling
Background: The RNDIS query handler did not validate `InformationBufferOffset` and `InformationBufferLength`, which could result in buffer overflows due to unchecked offset calculations exceeding the message length.
Impact: An attacker could exploit this flaw to trigger memory corruption or crashes in the USB Net device emulator, potentially leading to denial of service or arbitrary code execution.
Fix: Added checks to ensure the combined buffer offset and length do not exceed the total message length, avoiding unsafe memory access.
'''

# 定义一个函数来分析代码并返回结果
def analyze_code(code, commit_msg):
    code = code[:10000]
    messages = [
        {
            "role": "user",
            "content": "Given a piece of code and original commit message, rewrite the commit message in a clear, standardized format within 80 words."
                       "Desired Format:"
                       "Summary:"
                       "<Module>: <One-line fix summary>"
                       "Background:"
                       "<Briefly explain the root cause>"
                       "Impact:"
                       "<Describe the security impact if exploited>"
                       "Fix:"
                       "<Explain how this patch resolves the problem.>"
                       "Example:"
                       "Code Snippet: {}" 
                       "Commit Message: {}"
                       "Output: {}"
                       "Code Snippet: {}"
                       "Commit Message: {}".format(example_code, example_commit,example_out, code, commit_msg),
        }
    ]

    response = client.chat.completions.create(
        messages=messages,
        model="gpt-3.5-turbo",
        temperature=0.1,
        top_p=1,
    )
    for message in response.choices:
        print(message.message.content)
    result = {}
    for message in response.choices:
        for line in message.message.content.split('\n'):
            if line.startswith('Summary:'):
                result['Summary'] = line[len('Summary: '):].strip()
            elif line.startswith('Background:'):
                result['Background'] = line[len('Background: '):].strip()
            elif line.startswith('Impact:'):
                result['Impact'] = line[len('Impact: '):].strip()
            elif line.startswith('Fix:'):
                result['Fix'] = line[len('Fix: '):].strip()
    return result


# 创建一个新的DataFrame来存储结果
results_df = pd.DataFrame(columns=['cwe_ids', 'Summary', 'Background', 'Impact', 'Fix'])
count = 0
start_index = 0
csv_filename = 'MegaVul_new_com_msg.xlsx'

# 遍历DataFrame，分析代码并将结果逐条写入Excel文件
for index, row in df.iloc[start_index:].iterrows():
    result = analyze_code(row['func_before_recom'], row['commit_msg'])
    new_row = [
        row['cwe_ids'],
        result.get('Summary', ''),
        result.get('Background', ''),
        result.get('Impact', ''),
        result.get('Fix', ''),
    ]
    # 如果文件不存在，创建一个新的工作簿并写入表头和数据
    if not os.path.exists(csv_filename):
        wb = Workbook()
        ws = wb.active
        # 写入表头
        ws.append(['cwe_ids', 'Summary', 'Background', 'Impact', 'Fix'])
        # 写入数据
        ws.append(new_row)
        wb.save(csv_filename)
    else:
        # 如果文件存在，追加数据
        wb = load_workbook(csv_filename)
        ws = wb.active
        ws.append(new_row)
        wb.save(csv_filename)
    count += 1
    print("处理完成{}个啦".format(count))